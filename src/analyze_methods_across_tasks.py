#!/usr/bin/env python3
"""
Script to analyze and compare different methods (bestfs, mcts, tot_bfs, explorer_v3) 
across different tasks (Countdown and Sudoku variants).
"""

import argparse
import os
import pickle
import matplotlib.pyplot as plt
import seaborn as sns
import numpy as np
import pandas as pd
from tqdm import tqdm
import statsmodels.stats.proportion as smp
from scipy.stats import norm, t
from itertools import combinations
import statsmodels.api as sm
from scipy import stats
from scipy.stats import binomtest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from src.utils.common_utils import get_countdown_output_path, get_sudoku_output_path
from llm_mcts import MonteCarloTreeSearchNode


def parse_args():
    parser = argparse.ArgumentParser(description="Analyze methods across different tasks")
    
    # Basic parameters
    parser.add_argument("--countdown_data_dir", type=str, required=True, 
                       help="Directory with Countdown data")
    parser.add_argument("--sudoku_data_dir", type=str, required=True, 
                       help="Directory with Sudoku data")
    parser.add_argument("--output_dir", type=str, required=True, 
                       help="Directory to save analysis results")
    parser.add_argument("--save_figs", action="store_true",
                       help="Save figures to output directory")
    
    # Batch parameters
    parser.add_argument("--batch_nums", type=str, default="1",
                       help="Comma-separated list of batch numbers to analyze")
    parser.add_argument("--batch_size", type=int, default=1,
                       help="Size of each batch")
    
    # Model parameters
    parser.add_argument("--model_names", type=str, default="gpt-4o",
                       help="Comma-separated list of model names used")
    parser.add_argument("--temperature", type=float, default=0.0, 
                       help="Temperature used for generation")
    
    # Method parameters
    parser.add_argument("--methods", type=str, default="tot_bfs,bestfs,mcts,explorer_v3",
                       help="Comma-separated list of methods to include in plots")
    
    # MCTS-specific parameters
    parser.add_argument("--c_values", type=str, default="0.5",
                       help="Comma-separated list of c-values for MCTS")
    
    return parser.parse_args()


def load_raw_data(args, task_config):
    """
    Load raw data for a specific task configuration across all batches.
    
    Args:
        args: Command line arguments
        task_config: Dictionary containing task configuration
            For Countdown: {"type": "countdown", "split": "val", "game_diff": int, "mtu": int}
            For Sudoku: {"type": "sudoku", "size": int, "width": int, "height": int, "difficulty": str, "mtu": int}
    
    Returns:
        dict: Dictionary mapping method_model keys to their raw data
    """
    methods = args.methods.split(",")
    model_names = args.model_names.split(",")
    c_values = [float(x) for x in args.c_values.split(",")]
    batch_nums = [int(x) for x in args.batch_nums.split(",")]
    
    # Dictionary to store results
    raw_data = {}
    
    # Select appropriate data directory based on task type
    data_dir = args.countdown_data_dir if task_config["type"] == "countdown" else args.sudoku_data_dir
    
    for model_name in model_names:
        for method in methods:
            # Handle MCTS variants
            if method == "mcts":
                for c_value in c_values:
                    method_key = f"mcts(c={c_value})"
                    method_model_key = f"{method_key}_{model_name}"
                    
                    # List to store data from all batches
                    all_batch_data = []
                    
                    # Load data from each batch
                    for batch_idx in batch_nums:
                        # Get pickle path based on task type
                        if task_config["type"] == "countdown":
                            pickle_path = get_countdown_output_path(
                                data_dir,
                                "mcts",
                                batch_idx,
                                args.batch_size,
                                task_config["split"],
                                task_config["game_diff"],
                                model_name,
                                args.temperature,
                                max_token_usage=task_config["mtu"],
                                c_value=c_value
                            )
                        else:  # sudoku
                            pickle_path = get_sudoku_output_path(
                                data_dir,
                                "mcts",
                                batch_idx,
                                args.batch_size,
                                task_config["size"],
                                task_config["difficulty"],
                                model_name,
                                args.temperature,
                                max_token_usage=task_config["mtu"],
                                width=task_config["width"],
                                height=task_config["height"],
                                c_value=c_value
                            )
                        
                        try:
                            with open(pickle_path, 'rb') as f:
                                print(f"Loading {pickle_path}")
                                batch_data = pickle.load(f)
                                all_batch_data.extend(batch_data)
                        except (FileNotFoundError, EOFError) as e:
                            print(f"Error loading {pickle_path}: {e}")
                            continue
                    
                    if all_batch_data:
                        raw_data[method_model_key] = all_batch_data
            else:
                method_model_key = f"{method}_{model_name}"
                
                # List to store data from all batches
                all_batch_data = []
                
                # Load data from each batch
                for batch_idx in batch_nums:
                    # Get pickle path based on task type
                    if task_config["type"] == "countdown":
                        pickle_path = get_countdown_output_path(
                            data_dir,
                            method,
                            batch_idx,
                            args.batch_size,
                            task_config["split"],
                            task_config["game_diff"],
                            model_name,
                            args.temperature,
                            max_token_usage=task_config["mtu"]
                        )
                    else:  # sudoku
                        pickle_path = get_sudoku_output_path(
                            data_dir,
                            method,
                            batch_idx,
                            args.batch_size,
                            task_config["size"],
                            task_config["difficulty"],
                            model_name,
                            args.temperature,
                            max_token_usage=task_config["mtu"],
                            width=task_config["width"],
                            height=task_config["height"]
                        )
                    
                    try:
                        with open(pickle_path, 'rb') as f:
                            print(f"Loading {pickle_path}")
                            batch_data = pickle.load(f)
                            all_batch_data.extend(batch_data)
                    except (FileNotFoundError, EOFError) as e:
                        print(f"Error loading {pickle_path}: {e}")
                        continue
                
                if all_batch_data:
                    raw_data[method_model_key] = all_batch_data
    
    return raw_data

def extract_metrics_from_raw_data(raw_data, method_name, model_name, task_name, args):
    """
    Extract metrics from raw data, handling multiple iterations per game.
    Focuses on win rates, performance, and token usage.
    
    Args:
        raw_data: List of game outputs from pickle files
        method_name: Name of the method
        model_name: Name of the model
        task_name: Name of the task
        args: Command line arguments
    
    Returns:
        dict: Dictionary containing extracted metrics with averages and standard deviations
    """
    if not raw_data:
        return None
    
    # Calculate metrics across all games
    total_games = len(raw_data)
    
    # Lists to store metrics for each game
    game_win_rates = []
    game_performances = []
    game_token_usages = []
    
    # Lists to store confidence intervals for each game
    game_confidence_intervals = []
    
    # Lists to store metrics for winning games only
    winning_game_token_usages = []
    
    # Process each game
    for game in raw_data:
        # Calculate win rate statistics
        game_results_list = [1 if attempt["won"] else 0 for attempt in game]
        wins = sum(game_results_list)
        n = len(game_results_list)
        
        if n > 0:
            # Calculate confidence interval for this game's win rate
            try:
                ci_low, ci_upp = smp.proportion_confint(wins, n, method='wilson')
                game_confidence_intervals.append((ci_low, ci_upp))
            except Exception as e:
                print(f"Warning: Could not calculate confidence interval: {e}")
                game_confidence_intervals.append((None, None))
        else:
            print("Warning: No attempts found in game")
            game_confidence_intervals.append((None, None))
        
        game_wins = sum(1 for attempt in game if attempt["won"])
        game_win_rate = game_wins / len(game)
        game_win_rates.append(game_win_rate)
        
        # Calculate performance based on the selected metric type
        if args.performance_metric == "threshold":
            game_performance = 1.0 if game_win_rate > 0.5 else 0.0
        else:  # statistical
            ci_low, ci_upp = game_confidence_intervals[-1]
            game_performance = 1.0 if ci_low is not None and ci_low > 0.5 else 0.0
        
        game_performances.append(game_performance)
        
        # Calculate average token usage for this game
        game_tokens = []
        for attempt in game:
            if "total_tokens" in attempt:
                game_tokens.append(attempt["total_tokens"])
            elif "token_usage" in attempt:
                tokens = sum(usage["total_tokens"] for _, usage in attempt["token_usage"] if usage and "total_tokens" in usage)
                game_tokens.append(tokens)
        
        # Calculate average token usage for this game
        avg_game_tokens = np.mean(game_tokens)
        game_token_usages.append(avg_game_tokens)
        
        # If this is a winning game (based on the selected metric), store token usage
        if game_performance > 0:
            winning_game_token_usages.append(avg_game_tokens)
    
    # Calculate overall metrics
    overall_win_rate = np.mean(game_win_rates)
    overall_performance = np.mean(game_performances)
    overall_tokens = np.mean(game_token_usages)
    
    # Calculate overall performance confidence interval using Wilson's method
    total_performance_wins = sum(game_performances)
    total_performance_attempts = len(game_performances)
    try:
        overall_performance_ci = smp.proportion_confint(total_performance_wins, total_performance_attempts, method='wilson')
    except Exception as e:
        print(f"Warning: Could not calculate overall performance confidence interval: {e}")
        overall_performance_ci = (None, None)
    
    # Calculate standard deviations
    win_rate_std = np.std(game_win_rates)
    performance_std = np.std(game_performances)
    tokens_std = np.std(game_token_usages)
    
    # Calculate winning game metrics
    winning_tokens = np.mean(winning_game_token_usages) if winning_game_token_usages else 0
    winning_tokens_std = np.std(winning_game_token_usages) if winning_game_token_usages else 0

    # Print overall performance with confidence interval
    print(f"\n📊 {task_name} - {method_name} ({model_name})")
    print(f"  Overall performance: {overall_performance:.2f}")
    if overall_performance_ci[0] is not None:
        print(f"  95% CI: ({overall_performance_ci[0]:.2f}, {overall_performance_ci[1]:.2f})")
    print(f"  Overall win rate: {overall_win_rate:.2f}")

    metrics = {
        "method": method_name,
        "model": model_name,
        "total_games": total_games,
        "win_rate": overall_win_rate,
        "win_rate_std": win_rate_std,
        "performance": overall_performance,
        "performance_std": performance_std,
        "avg_tokens": overall_tokens,
        "tokens_std": tokens_std,
        "winning_tokens": winning_tokens,
        "winning_tokens_std": winning_tokens_std,
        # Add confidence interval information
        "overall_performance_ci": overall_performance_ci,
        "game_confidence_intervals": game_confidence_intervals,
        "game_win_rates": game_win_rates,
        "game_performances": game_performances,
        "game_sample_sizes": [len(game) for game in raw_data],
        "game_token_usages": game_token_usages,
        "winning_game_token_usages": winning_game_token_usages,
        # Store raw data for iteration-level analysis
        "raw_data": raw_data
    }
    
    return metrics



def extract_metrics_from_combined_data(combined_data, task_name, args):
    """
    Extract metrics from combined raw data.
    
    Args:
        combined_data: Dictionary mapping method_model keys to their combined raw data
        task_name: Name of the task
        args: Command line arguments
    
    Returns:
        dict: Dictionary mapping method_model keys to their metrics
    """
    results = {}
    
    for method_model_key, raw_data in combined_data.items():
        # Extract method name and model name from the key
        if "mcts" in method_model_key:
            method_name = method_model_key.split("_")[0]  # Includes c value
            model_name = "_".join(method_model_key.split("_")[1:])
        else:
            method_name = method_model_key.split("_")[0]
            model_name = "_".join(method_model_key.split("_")[1:])
        
        results[method_model_key] = extract_metrics_from_raw_data(raw_data, method_name, model_name, task_name, args)
    
    return results




def get_method_display_name(method_name):
    """
    Map method names to their display names in plots.
    """
    # Handle MCTS with c-value
    if "mcts" in method_name:
        # Extract the c-value from the method name
        c_value = method_name.split("c=")[1].split(")")[0]
        return f"MCTS(c={c_value})"
    
    # Handle other methods
    display_names = {
        "explorer_v3": "LFS",
        "explorer": "LFS",  # Add mapping for 'explorer' to display as 'LFS'
        "tot_bfs": "ToT-BFS",
        "tot": "ToT-BFS",  # Add mapping for 'tot' to match color dictionary
        "mcts": "MCTS",
        "bestfs": "BESTFS"
    }
    return display_names.get(method_name, method_name)

def get_method_color_key(method_name):
    """
    Get the key to use for looking up colors in the method_colors dictionary.
    """
    # Handle MCTS with c-value
    if "mcts" in method_name:
        if "c=" in method_name:
            return method_name  # Already in correct format
        else:
            return f"mcts(c=1.0)"  # Default c-value if not specified
    
    # Handle other methods
    color_keys = {
        "tot": "tot_bfs",  # Map 'tot' to 'tot_bfs' for color lookup
        "explorer": "explorer_v3",  # Map 'explorer' to 'explorer_v3' for color lookup
    }
    return color_keys.get(method_name, method_name)

def plot_task_metrics(results, task_name, args, task=None, method_colors=None):
    """
    Create separate plots for win rates and token usage.
    """
    if not results:
        print(f"No data available for task {task_name}")
        return
    
    # Set global font sizes
    plt.rcParams.update({
        'font.size': 18,
        'axes.titlesize': 24,
        'axes.labelsize': 20,
        'xtick.labelsize': 18,
        'ytick.labelsize': 18,
        'legend.fontsize': 18,
        'font.weight': 'normal',
        'axes.labelweight': 'normal',
        'axes.titleweight': 'normal'
    })
    
    # Prepare data for plots
    methods = []
    original_methods = []  # Store original method names for color lookup
    win_rates = []
    win_rate_cis = []
    token_usages = []
    token_cis = []
    
    # Get MTU value from task object
    mtu = task["mtu"] if task else 1000000  # Default MTU if task not provided
    
    for method_model_key, metrics in results.items():
        if metrics is None:
            continue
            
        # Get display name for the method
        method_name = metrics["method"]
        if "mcts" in method_name:
            display_name = get_method_display_name(method_name.split("_")[0])
            original_method = method_name.split("_")[0]
            color_key = original_method
        else:
            display_name = get_method_display_name(method_name)
            original_method = method_name
            color_key = get_method_color_key(original_method)
        
        methods.append(display_name)
        original_methods.append(original_method)
        
        # Calculate win rate CI
        win_rate = metrics["win_rate"]
        if win_rate == 0 or win_rate == 1:
            win_rate_ci = (win_rate * 100, win_rate * 100)
        else:
            win_rate_std = np.sqrt(win_rate * (1 - win_rate))
            win_rate_ci = t.interval(0.95, df=metrics["total_games"]-1, 
                                    loc=win_rate, 
                                    scale=win_rate_std/np.sqrt(metrics["total_games"]))
            win_rate_ci = (max(0, win_rate_ci[0]) * 100, min(1, win_rate_ci[1]) * 100)
        win_rates.append(win_rate * 100)
        win_rate_cis.append(win_rate_ci)
        
        # Calculate token usage CI
        tokens = metrics["avg_tokens"]
        token_std = metrics["tokens_std"]
        if token_std > 0 and metrics["total_games"] > 1:
            token_ci = t.interval(0.95, df=metrics["total_games"]-1, 
                                 loc=tokens, 
                                 scale=token_std/np.sqrt(metrics["total_games"]))
            token_ci = (max(0, token_ci[0]), min(mtu, token_ci[1]))
        else:
            token_ci = (tokens, tokens)
        token_usages.append(tokens)
        token_cis.append(token_ci)
    
    # Convert lists to numpy arrays
    win_rates = np.array(win_rates)
    token_usages = np.array(token_usages)
    
    # Plot win rates
    plt.figure(figsize=(15, 8))
    x = np.arange(len(methods))
    width = 0.8
    
    for i, (method, original_method, win_rate, win_ci) in enumerate(zip(methods, original_methods, win_rates, win_rate_cis)):
        color_key = get_method_color_key(original_method) if "mcts" not in original_method else original_method
        color = method_colors[color_key] if method_colors else None
        plt.bar(x[i], win_rate, width, color=color, label=method)
        plt.text(x[i], win_rate + 2, f'*{win_rate:.1f}%*\n[{win_ci[0]:.1f},\n{win_ci[1]:.1f}]', 
                ha='center', va='bottom', fontsize=16)
    
    plt.xlabel('Methods', fontsize=20)
    plt.ylabel('Win Rate (%)', fontsize=20)
    plt.title(f'Win Rates for {task_name}\nModel: {args.model_names}', fontsize=24, pad=20)
    plt.xticks(x, methods, rotation=45, ha='right')
    plt.grid(True, linestyle='--', alpha=0.3)
    plt.legend(fontsize=18)
    
    # Set y-axis limits with padding for labels and legend
    max_win_rate = np.max([ci[1] for ci in win_rate_cis])  # Use upper CI bound for max value
    plt.ylim(0, max_win_rate * 1.5)  # Add 50% padding for labels and legend
    
    plt.tight_layout()
    
    if args.save_figs:
        safe_task_name = task_name.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
        safe_model_name = args.model_names.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
        safe_c_value = args.c_values.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
        safe_methods = args.methods.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
        filename = os.path.join(args.output_dir, f"{args.analysis_type}_{safe_task_name}_{safe_model_name}_c{safe_c_value}_methods_{safe_methods}_win_rates.pdf")
        plt.savefig(filename, bbox_inches='tight', dpi=300)
        plt.close()
    else:
        plt.show()
        plt.close()
    
    # Plot token usage
    plt.figure(figsize=(15, 8))
    
    for i, (method, original_method, tokens, token_ci) in enumerate(zip(methods, original_methods, token_usages, token_cis)):
        color_key = get_method_color_key(original_method) if "mcts" not in original_method else original_method
        color = method_colors[color_key] if method_colors else None
        plt.bar(x[i], tokens, width, color=color, label=method)
        plt.text(x[i], tokens + max(token_usages) * 0.02, f'*{tokens:.0f}*\n[{token_ci[0]:.0f},\n{token_ci[1]:.0f}]', 
                ha='center', va='bottom', fontsize=16)
    
    plt.xlabel('Methods', fontsize=20)
    plt.ylabel('Token Usage', fontsize=20)
    plt.title(f'Token Usage for {task_name}\nModel: {args.model_names}', fontsize=24, pad=20)
    plt.xticks(x, methods, rotation=45, ha='right')
    plt.grid(True, linestyle='--', alpha=0.3)
    plt.legend(fontsize=18)
    
    # Set y-axis limits with padding for labels and legend
    max_tokens = np.max([ci[1] for ci in token_cis])  # Use upper CI bound for max value
    plt.ylim(0, max_tokens * 1.7)  # Increased padding to 70% for labels and legend
    
    plt.tight_layout()
    
    if args.save_figs:
        safe_task_name = task_name.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
        safe_model_name = args.model_names.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
        safe_c_value = args.c_values.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
        safe_methods = args.methods.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
        filename = os.path.join(args.output_dir, f"{args.analysis_type}_{safe_task_name}_{safe_model_name}_c{safe_c_value}_methods_{safe_methods}_token_usage.pdf")
        plt.savefig(filename, bbox_inches='tight', dpi=300)
        plt.close()
    else:
        plt.show()
        plt.close()
    
    # Create win rate vs token usage scatter plot
    plt.figure(figsize=(15, 8))
    for i, (method, original_method, tokens, win_rate) in enumerate(zip(methods, original_methods, token_usages, win_rates)):
        color_key = get_method_color_key(original_method) if "mcts" not in original_method else original_method
        color = method_colors[color_key] if method_colors else None
        plt.scatter(tokens, win_rate, label=method, s=100, color=color)
    plt.xlabel('Token Usage', fontsize=20)
    plt.ylabel('Win Rate (%)', fontsize=20)
    plt.title(f'Win Rate vs Token Usage for {task_name}\nModel: {args.model_names}', fontsize=24, pad=20)
    plt.legend(frameon=True, facecolor='white', edgecolor='black', fontsize=18)
    plt.grid(True, linestyle='--', alpha=0.3)
    plt.tight_layout()
    
    if args.save_figs:
        safe_task_name = task_name.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
        safe_model_name = args.model_names.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
        safe_c_value = args.c_values.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
        safe_methods = args.methods.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
        filename = os.path.join(args.output_dir, f"{args.analysis_type}_{safe_task_name}_{safe_model_name}_c{safe_c_value}_methods_{safe_methods}_win_rate_vs_tokens.pdf")
        plt.savefig(filename, bbox_inches='tight', dpi=300)
        plt.close()
    else:
        plt.show()
        plt.close()


def plot_performance_profiles(results_by_task, args, method_colors):
    """
    Plot performance profiles across tasks for win rates.
    """
    # Set global font sizes
    plt.rcParams.update({
        'font.size': 18,  # Increased base font size
        'axes.titlesize': 24,  # Increased title size
        'axes.labelsize': 20,  # Increased label size
        'xtick.labelsize': 18,  # Increased tick label size
        'ytick.labelsize': 18,  # Increased tick label size
        'legend.fontsize': 18,  # Increased legend font size
        'font.weight': 'normal',
        'axes.labelweight': 'normal',
        'axes.titleweight': 'normal'
    })
    
    # Define different markers for each method
    markers = ['o', 's', '^', 'D', '*', 'v', '>', '<', 'p', 'h']
    
    # Define metrics to plot
    metric_configs = [
        ("win_rate_token_ratio", "Win Rate / Token Usage"),
        ("win_rate", "Win Rate")
    ]
    
    # Get methods in the order specified by command line arguments
    methods = args.methods.split(",")
    model_names = args.model_names.split(",")
    c_values = [float(x) for x in args.c_values.split(",")]
    
    # Create ordered list of method_model keys
    ordered_method_keys = []
    for method in methods:
        if method == "mcts":
            for c_value in c_values:
                for model_name in model_names:
                    method_key = f"mcts(c={c_value})"
                    method_model_key = f"{method_key}_{model_name}"
                    ordered_method_keys.append(method_model_key)
        else:
            for model_name in model_names:
                method_model_key = f"{method}_{model_name}"
                ordered_method_keys.append(method_model_key)
    
    # Calculate performance ratios for each metric
    method_ratios = {}
    for metric_name, metric_display in metric_configs:
        method_ratios[metric_name] = {}
        
        # First pass: find best performance for each task
        best_tokens = {}
        best_win_rates = {}
        
        for task_name, task_results in results_by_task.items():
            best_tokens[task_name] = float('inf')
            best_win_rates[task_name] = float('-inf')
            
            # Find best values for each metric
            for method_model_key, metrics in task_results.items():
                if metrics is None:
                    print(f"Warning: No metrics available for {method_model_key} on task {task_name}")
                    continue
                
                best_tokens[task_name] = min(best_tokens[task_name], metrics["avg_tokens"])
                best_win_rates[task_name] = max(best_win_rates[task_name], metrics["win_rate"])
        
        # Second pass: calculate ratios
        for method_model_key in ordered_method_keys:
            ratios = []
            
            for task_name, task_results in results_by_task.items():
                metrics = task_results.get(method_model_key)
                
                # If metrics is None, assign infinite ratio (worst possible performance)
                if metrics is None:
                    ratios.append(float('inf'))
                    continue
                
                if metric_name == "win_rate_token_ratio":
                    # Calculate efficiency ratio: (best_win_rate/win_rate) / (tokens/best_tokens)
                    win_rate_ratio = best_win_rates[task_name] / metrics["win_rate"] if metrics["win_rate"] > 0 else float('inf')
                    token_ratio = metrics["avg_tokens"] / best_tokens[task_name] if best_tokens[task_name] > 0 else float('inf')
                    ratio = win_rate_ratio * token_ratio
                
                else:  # win_rate
                    ratio = best_win_rates[task_name] / metrics["win_rate"] if metrics["win_rate"] > 0 else float('inf')
                
                ratios.append(ratio)
            
            if ratios:
                method_ratios[metric_name][method_model_key] = ratios
    
    # Plot performance profiles for each metric
    # Use more granular tau values and a smaller range for better visibility
    tau_values = np.logspace(0, np.log10(10), 200)  # 1 to 10 on log scale with more points
    
    # Create a single figure for win rate
    plt.figure(figsize=(12, 8))
    
    # Plot each method in the specified order
    for i, method_model_key in enumerate(ordered_method_keys):
        # Extract base method name for color lookup and display name
        if "mcts" in method_model_key:
            base_method = method_model_key.split("_")[0]  # Includes c value
            display_name = get_method_display_name(base_method)  # Just use the MCTS part with c value
            color_key = base_method  # MCTS methods are already in correct format
        else:
            base_method = method_model_key.split("_")[0]
            display_name = get_method_display_name(base_method)
            color_key = get_method_color_key(base_method)  # Get correct color key
        
        color = method_colors[color_key]
        marker = markers[i % len(markers)]  # Cycle through markers if more methods than markers
        
        # Calculate profile for win rate
        win_rate_profile = []
        
        for tau in tau_values:
            # Win rate profile
            win_rate_ratios = method_ratios["win_rate"][method_model_key]
            win_rate_count = sum(1 for r in win_rate_ratios if r <= tau)
            win_rate_profile.append(win_rate_count / len(win_rate_ratios))
        
        # Calculate AUC for win rate
        win_rate_auc = np.trapz(win_rate_profile, tau_values)
        
        # Plot with thicker lines and markers
        plt.plot(tau_values, win_rate_profile, 
                label=f"{display_name} (AUP: {win_rate_auc:.3f})", 
                color=color, 
                linewidth=3.5,  # Increased from 2.5 to 3.5
                marker=marker,  # Use the assigned marker
                markersize=10,  # Increased marker size
                markevery=20)  # Add markers every 20 points
    
    plt.xscale('log')  # Set x-axis to logarithmic scale
    plt.xlabel('Performance Ratio (τ)', fontsize=20)
    plt.ylabel('Fraction of Tasks', fontsize=20)
    plt.title(f'Win Rate Profile\nModel: {args.model_names}', fontsize=24, pad=20)
    plt.grid(True, linestyle='--', alpha=0.3)
    
    # Let matplotlib choose the best legend location
    plt.legend(frameon=True, facecolor='white', edgecolor='black', fontsize=18)
    
    # Add grid lines for better readability
    plt.grid(True, which='both', linestyle='--', alpha=0.3)
    
    # Add minor grid lines
    plt.minorticks_on()
    plt.grid(True, which='minor', linestyle=':', alpha=0.2)
    
    plt.tight_layout()
    
    if args.save_figs:
        # Create a safe filename by replacing spaces and special characters
        safe_model_name = args.model_names.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
        safe_c_value = args.c_values.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
        safe_methods = args.methods.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
        filename = os.path.join(args.output_dir, f"{args.analysis_type}_{safe_model_name}_c{safe_c_value}_methods_{safe_methods}_win_rate_profile.pdf")
        plt.savefig(filename, bbox_inches='tight')
        plt.close()
    else:
        plt.show()
        plt.close()
    
    # Plot token ratio metrics separately
    for metric_name, metric_display in metric_configs[:1]:  # Only plot token ratio metrics
        plt.figure(figsize=(12, 8))
        
        # Plot each method in the specified order
        for i, method_model_key in enumerate(ordered_method_keys):
            # Extract base method name for color lookup and display name
            if "mcts" in method_model_key:
                base_method = method_model_key.split("_")[0]  # Includes c value
                display_name = get_method_display_name(base_method)
                color_key = base_method  # MCTS methods are already in correct format
            else:
                base_method = method_model_key.split("_")[0]
                display_name = get_method_display_name(base_method)
                color_key = get_method_color_key(base_method)  # Get correct color key
            
            color = method_colors[color_key]
            marker = markers[i % len(markers)]  # Cycle through markers if more methods than markers
            
            # Calculate profile
            ratios = method_ratios[metric_name][method_model_key]
            profile = []
            for tau in tau_values:
                count = sum(1 for r in ratios if r <= tau)
                profile.append(count / len(ratios))
            
            # Calculate area under curve (AUC)
            auc = np.trapz(profile, tau_values)
            
            # Plot with thicker lines and markers
            plt.plot(tau_values, profile, 
                    label=f"{display_name} (AUP: {auc:.3f})", 
                    color=color, 
                    linewidth=3.5,  # Increased from 2.5 to 3.5
                    marker=marker,  # Use the assigned marker
                    markersize=10,  # Increased marker size
                    markevery=20)  # Add markers every 20 points)
        
        plt.xscale('log')  # Set x-axis to logarithmic scale
        plt.xlabel('Performance Ratio (τ)', fontsize=20)
        plt.ylabel('Fraction of Tasks', fontsize=20)
        plt.title(f'{metric_display} Profile\nModel: {args.model_names}', fontsize=24, pad=20)
        plt.grid(True, linestyle='--', alpha=0.3)
        
        # Let matplotlib choose the best legend location
        plt.legend(frameon=True, facecolor='white', edgecolor='black', fontsize=18)
        
        # Add grid lines for better readability
        plt.grid(True, which='both', linestyle='--', alpha=0.3)
        
        # Add minor grid lines
        plt.minorticks_on()
        plt.grid(True, which='minor', linestyle=':', alpha=0.2)
        
        plt.tight_layout()
        
        if args.save_figs:
            # Create a safe filename by replacing spaces and special characters
            safe_model_name = args.model_names.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
            safe_c_value = args.c_values.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
            safe_methods = args.methods.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
            filename = os.path.join(args.output_dir, f"{args.analysis_type}_{safe_model_name}_c{safe_c_value}_methods_{safe_methods}_performance_profile_{metric_name}.pdf")
            plt.savefig(filename, bbox_inches='tight')
            plt.close()
        else:
            plt.show()
            plt.close()


def plot_cumulative_wins_vs_tokens(results_by_task, args, method_colors):
    """
    Plot cumulative wins vs token usage for winning iterations, with separate plots for Countdown and Sudoku.
    Methods that stop winning will have their lines extended horizontally.
    Uses logarithmic scale for token usage to better visualize distribution across different ranges.
    
    Args:
        results_by_task: Dictionary mapping task names to their results
        args: Command line arguments
        method_colors: Dictionary mapping method names to their colors
    """
    # Set global font sizes
    plt.rcParams.update({
        'font.size': 32,  # Increased from 24 to 32
        'axes.titlesize': 40,  # Increased from 32 to 40
        'axes.labelsize': 36,  # Increased from 28 to 36
        'xtick.labelsize': 32,  # Increased from 24 to 32
        'ytick.labelsize': 32,  # Increased from 24 to 32
        'legend.fontsize': 32,  # Increased from 24 to 32
        'font.weight': 'normal',
        'axes.labelweight': 'normal',
        'axes.titleweight': 'normal'
    })
    
    # Define MTU limits for each game type
    mtu_limits = {
        "Countdown": 1000000,
        "Sudoku": 500000
    }
    
    # Separate tasks into Countdown and Sudoku
    countdown_tasks = [task_name for task_name in results_by_task.keys() if "Countdown" in task_name]
    sudoku_tasks = [task_name for task_name in results_by_task.keys() if "Sudoku" in task_name]
    
    # Get methods in the order specified by command line arguments
    methods = args.methods.split(",")
    model_names = args.model_names.split(",")
    c_values = [float(x) for x in args.c_values.split(",")]
    
    # Create ordered list of method_model keys
    ordered_method_keys = []
    for method in methods:
        if method == "mcts":
            for c_value in c_values:
                for model_name in model_names:
                    method_key = f"mcts(c={c_value})"
                    method_model_key = f"{method_key}_{model_name}"
                    ordered_method_keys.append(method_model_key)
        else:
            for model_name in model_names:
                method_model_key = f"{method}_{model_name}"
                ordered_method_keys.append(method_model_key)
    
    # Create separate plots for Countdown and Sudoku
    for game_type, tasks in [("Countdown", countdown_tasks), ("Sudoku", sudoku_tasks)]:
        # Create a square figure
        plt.figure(figsize=(15, 12))  # Changed from (15, 8) to (12, 12) for square aspect ratio
        
        # First pass: find max tokens across all methods
        max_tokens = 0
        min_tokens = float('inf')  # Track minimum tokens for log scale
        for method_model_key in ordered_method_keys:
            for task_name in tasks:
                if task_name in results_by_task and method_model_key in results_by_task[task_name]:
                    metrics = results_by_task[task_name][method_model_key]
                    if metrics is not None and "raw_data" in metrics:
                        for game in metrics["raw_data"]:
                            for attempt in game:
                                if attempt["won"]:
                                    if "total_tokens" in attempt:
                                        max_tokens = max(max_tokens, min(attempt["total_tokens"], mtu_limits[game_type]))
                                        min_tokens = min(min_tokens, attempt["total_tokens"])
                                    elif "token_usage" in attempt:
                                        tokens = sum(usage["total_tokens"] for _, usage in attempt["token_usage"] if usage and "total_tokens" in usage)
                                        max_tokens = max(max_tokens, min(tokens, mtu_limits[game_type]))
                                        min_tokens = min(min_tokens, tokens)
        
        # Add padding for log scale
        max_tokens = min(int(max_tokens * 1.1), mtu_limits[game_type])  # Ensure we don't exceed MTU
        min_tokens = max(1, int(min_tokens * 0.9))  # Ensure minimum is at least 1 for log scale
        
        # Define different markers for each method
        markers = ['o', 's', '^', 'D', '*', 'v', '>', '<', 'p', 'h']
        
        # Process each method
        for i, method_model_key in enumerate(ordered_method_keys):
            # Extract base method name for color lookup and display name
            if "mcts" in method_model_key:
                base_method = method_model_key.split("_")[0]  # Includes c value
                display_name = get_method_display_name(base_method)
                color_key = base_method  # MCTS methods are already in correct format
            else:
                base_method = method_model_key.split("_")[0]
                display_name = get_method_display_name(base_method)
                color_key = get_method_color_key(base_method)  # Get correct color key
            
            color = method_colors[color_key]
            marker = markers[i % len(markers)]  # Cycle through markers if more methods than markers
            
            # Collect all winning iterations and their token usage
            winning_tokens = []
            
            for task_name in tasks:
                if task_name in results_by_task and method_model_key in results_by_task[task_name]:
                    metrics = results_by_task[task_name][method_model_key]
                    if metrics is not None and "raw_data" in metrics:
                        for game in metrics["raw_data"]:
                            for attempt in game:
                                if attempt["won"]:
                                    # Get token usage
                                    if "total_tokens" in attempt:
                                        tokens = min(attempt["total_tokens"], mtu_limits[game_type])
                                    elif "token_usage" in attempt:
                                        tokens = min(sum(usage["total_tokens"] for _, usage in attempt["token_usage"] if usage and "total_tokens" in usage), mtu_limits[game_type])
                                    else:
                                        continue
                                    winning_tokens.append(tokens)
            
            if winning_tokens:
                # Sort by token usage
                winning_tokens.sort()
                
                # Create x and y values for plotting
                x_values = winning_tokens
                y_values = list(range(1, len(winning_tokens) + 1))
                
                # Add horizontal line extension if there are wins
                if x_values:
                    # Add the max tokens point to extend the line horizontally
                    x_values.append(max_tokens)
                    y_values.append(y_values[-1])  # Use the last y value
                
                # Plot the line with markers
                plt.plot(x_values, y_values, 
                        label=display_name,
                        color=color,
                        linewidth=5,  # Increased from 4 to 5
                        marker=marker,  # Use the assigned marker
                        markersize=18,  # Increased from 14 to 18
                        markevery=max(1, len(x_values)//20))  # Show markers every ~20 points
        
        # Customize the plot
        plt.xlabel('Token Usage (log scale)', fontsize=36)
        plt.ylabel('Cumulative Number of Wins', fontsize=36)
        plt.title(f'Cumulative Wins vs Token Usage for {game_type}\nModel: {args.model_names}', fontsize=40, pad=20)
        
        # Add grid lines
        plt.grid(True, which='major', linestyle='--', alpha=0.3)
        plt.grid(True, which='minor', linestyle=':', alpha=0.2)
        
        # Set x-axis to logarithmic scale
        plt.xscale('log')
        
        # Set x-axis limits dynamically
        plt.xlim(min_tokens, max_tokens)
        
        # Add minor grid lines for better readability
        plt.minorticks_on()
        
        # Format y-axis to show integer ticks
        plt.gca().yaxis.set_major_locator(plt.MaxNLocator(integer=True))
        
        # Add legend with white background and larger size
        legend = plt.legend(fontsize=32, 
                          frameon=True, 
                          facecolor='white', 
                          edgecolor='black',
                          loc='upper left',
                          bbox_to_anchor=(0.01, 0.99))
        
        # Add grid lines for better readability
        plt.grid(True, which='both', linestyle='--', alpha=0.3)
        
        # Add minor grid lines
        plt.minorticks_on()
        plt.grid(True, which='minor', linestyle=':', alpha=0.2)
        
        # Adjust layout to prevent legend overlap
        plt.tight_layout()
        
        if args.save_figs:
            # Create a safe filename
            safe_model_name = args.model_names.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
            safe_c_value = args.c_values.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
            safe_methods = args.methods.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
            filename = os.path.join(args.output_dir, f"{args.analysis_type}_{safe_model_name}_c{safe_c_value}_methods_{safe_methods}_cumulative_wins_{game_type.lower()}.pdf")
            plt.savefig(filename, bbox_inches='tight', dpi=300)  # Increased DPI for better quality
            plt.close()
        else:
            plt.show()
            plt.close()


def create_excel_summary(results_by_task, args, output_dir):
    """
    Create an Excel file summarizing win rates and AUP values across all tasks and methods.
    
    Args:
        results_by_task: Dictionary mapping task names to their results
        args: Command line arguments
        output_dir: Directory to save the Excel file
    """
    # Create a new Excel workbook
    wb = Workbook()
    
    # Create sheets
    win_rates_sheet = wb.active
    win_rates_sheet.title = "Win Rates"
    aup_sheet = wb.create_sheet("AUP Values")
    
    # Get methods in the order specified by command line arguments
    methods = args.methods.split(",")
    model_names = args.model_names.split(",")
    c_values = [float(x) for x in args.c_values.split(",")]
    
    # Create ordered list of method_model keys
    ordered_method_keys = []
    for method in methods:
        if method == "mcts":
            for c_value in c_values:
                for model_name in model_names:
                    method_key = f"mcts(c={c_value})"
                    method_model_key = f"{method_key}_{model_name}"
                    ordered_method_keys.append(method_model_key)
        else:
            for model_name in model_names:
                method_model_key = f"{method}_{model_name}"
                ordered_method_keys.append(method_model_key)
    
    # Prepare data for win rates
    win_rate_data = []
    for task_name, task_results in results_by_task.items():
        for method_model_key in ordered_method_keys:
            metrics = task_results.get(method_model_key)
            if metrics is not None:
                # Extract method name and model name
                if "mcts" in method_model_key:
                    method_name = method_model_key.split("_")[0]  # Includes c value
                    model_name = "_".join(method_model_key.split("_")[1:])
                else:
                    method_name = method_model_key.split("_")[0]
                    model_name = "_".join(method_model_key.split("_")[1:])
                
                win_rate_data.append({
                    'Task': task_name,
                    'Method': get_method_display_name(method_name),
                    'Model': model_name,
                    'Win Rate (%)': metrics['win_rate'] * 100,
                    'Total Games': metrics['total_games']
                })
    
    # Create win rates DataFrame and write to sheet
    win_rates_df = pd.DataFrame(win_rate_data)
    win_rates_df = win_rates_df.sort_values(['Task', 'Method'])
    
    # Write win rates data
    for col_idx, column in enumerate(win_rates_df.columns, 1):
        cell = win_rates_sheet.cell(row=1, column=col_idx)
        cell.value = column
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, row in enumerate(win_rates_df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = win_rates_sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='center')
    
    # Calculate AUP values
    tau_values = np.logspace(0, np.log10(10), 200)  # 1 to 10 on log scale
    aup_data = []
    
    # Calculate performance ratios for each metric
    method_ratios = {}
    for metric_name in ["win_rate", "win_rate_token_ratio"]:
        method_ratios[metric_name] = {}
        
        # First pass: find best performance for each task
        best_tokens = {}
        best_win_rates = {}
        
        for task_name, task_results in results_by_task.items():
            best_tokens[task_name] = float('inf')
            best_win_rates[task_name] = float('-inf')
            
            for method_model_key, metrics in task_results.items():
                if metrics is None:
                    continue
                
                best_tokens[task_name] = min(best_tokens[task_name], metrics["avg_tokens"])
                best_win_rates[task_name] = max(best_win_rates[task_name], metrics["win_rate"])
        
        # Second pass: calculate ratios and AUP
        for method_model_key in ordered_method_keys:
            ratios = []
            
            for task_name, task_results in results_by_task.items():
                metrics = task_results.get(method_model_key)
                
                if metrics is None:
                    ratios.append(float('inf'))
                    continue
                
                if metric_name == "win_rate_token_ratio":
                    win_rate_ratio = best_win_rates[task_name] / metrics["win_rate"] if metrics["win_rate"] > 0 else float('inf')
                    token_ratio = metrics["avg_tokens"] / best_tokens[task_name] if best_tokens[task_name] > 0 else float('inf')
                    ratio = win_rate_ratio * token_ratio
                else:  # win_rate
                    ratio = best_win_rates[task_name] / metrics["win_rate"] if metrics["win_rate"] > 0 else float('inf')
                
                ratios.append(ratio)
            
            if ratios:
                method_ratios[metric_name][method_model_key] = ratios
                
                # Calculate profile and AUP
                profile = []
                for tau in tau_values:
                    count = sum(1 for r in ratios if r <= tau)
                    profile.append(count / len(ratios))
                
                aup = np.trapz(profile, tau_values)
                
                # Extract method name and model name
                if "mcts" in method_model_key:
                    method_name = method_model_key.split("_")[0]
                    model_name = "_".join(method_model_key.split("_")[1:])
                else:
                    method_name = method_model_key.split("_")[0]
                    model_name = "_".join(method_model_key.split("_")[1:])
                
                aup_data.append({
                    'Method': get_method_display_name(method_name),
                    'Model': model_name,
                    'Metric': metric_name.replace('_', ' ').title(),
                    'AUP': aup
                })
    
    # Create AUP DataFrame and write to sheet
    aup_df = pd.DataFrame(aup_data)
    aup_df = aup_df.sort_values(['Metric', 'Method'])
    
    # Write AUP data
    for col_idx, column in enumerate(aup_df.columns, 1):
        cell = aup_sheet.cell(row=1, column=col_idx)
        cell.value = column
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, row in enumerate(aup_df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = aup_sheet.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.alignment = Alignment(horizontal='center')
    
    # Adjust column widths
    for sheet in [win_rates_sheet, aup_sheet]:
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
    
    # Save the workbook
    safe_model_name = args.model_names.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
    safe_c_value = args.c_values.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
    safe_methods = args.methods.replace(' ', '_').replace('(', '').replace(')', '').replace(',', '')
    filename = os.path.join(output_dir, f"{args.analysis_type}_{safe_model_name}_c{safe_c_value}_methods_{safe_methods}_summary.xlsx")
    wb.save(filename)
    print(f"\nExcel summary saved to: {filename}")


def create_fixed_method_colors():
    """
    Create a fixed color dictionary for all possible methods.
    This ensures consistent colors across all plots.
    """
    # Define all possible methods
    all_possible_methods = [
        "tot_bfs",
        "bestfs",
        "mcts",
        "explorer_v3",
        "mcts(c=0.5)",
        "mcts(c=1.0)",
        "mcts(c=2.5)"
    ]
    
    # Define the color palette
    colors = [
        '#AA4499',
        '#88CCEE',
        '#CC6677',
        '#44AA99',
        '#CC6677',
        '#AA4499',
        '#882255',
    ]
    
    # Create dictionary mapping methods to colors
    method_colors = {}
    for method, color in zip(all_possible_methods, colors):
        method_colors[method] = color
    
    return method_colors

def main():
    args = parse_args()
    
    # Create output directory if saving figures
    if args.save_figs:
        os.makedirs(args.output_dir, exist_ok=True)
    
    # Define tasks to analyze with their specific MTU values
    tasks = [
        # Countdown tasks
        {"type": "countdown", "split": "val", "game_diff": 3, "name": "Countdown (diff=3)", "mtu": 1000000},
        {"type": "countdown", "split": "val", "game_diff": 5, "name": "Countdown (diff=5)", "mtu": 1000000},
        {"type": "countdown", "split": "val", "game_diff": 7, "name": "Countdown (diff=7)", "mtu": 1000000},
        # Sudoku tasks
        {"type": "sudoku", "size": 4, "width": 2, "height": 2, "difficulty": "hard",
         "name": "Sudoku (width=2, height=2)", "mtu": 100000},
        {"type": "sudoku", "size": 6, "width": 2, "height": 3, "difficulty": "medium",
         "name": "Sudoku (width=2, height=3)", "mtu": 500000}
    ]
    
    # Store raw data for each task
    raw_data_by_task = {}
    
    # Load raw data for each task
    for task in tasks:
        print(f"\nLoading data for task: {task['name']}")
        raw_data = load_raw_data(args, task)
        raw_data_by_task[task["name"]] = raw_data
    
    # Create fixed color dictionary for all methods
    method_colors = create_fixed_method_colors()

    # Extract metrics for all tasks
    results_by_task = {task["name"]: extract_metrics_from_combined_data(raw_data_by_task[task["name"]], task["name"], args) 
                      for task in tasks}
    
    # Create Excel summary
    print("\nCreating Excel summary...")
    create_excel_summary(results_by_task, args, args.output_dir)
    
    # Plot cumulative wins vs token usage
    print("\nPlotting cumulative wins vs token usage...")
    plot_cumulative_wins_vs_tokens(results_by_task, args, method_colors)

    # Original per-task analysis
    for task in tasks:
        print(f"\nAnalyzing task: {task['name']}")
        results = extract_metrics_from_combined_data(raw_data_by_task[task["name"]], task["name"], args)
        plot_task_metrics(results, task["name"], args, task, method_colors)

    # Plot performance profiles across tasks
    print("\nPlotting performance profiles across tasks...")
    plot_performance_profiles(results_by_task, args, method_colors)
    



if __name__ == "__main__":
    main() 